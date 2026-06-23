"""
check_status.py – shows whether the monitor is alive, when it last ran,
how many changes have been captured, and whether any polls were missed.

Run it any time: python check_status.py
"""

import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

import config as cfg


def _age(iso: str) -> str:
    try:
        dt  = datetime.fromisoformat(iso)
        sec = int((datetime.now() - dt).total_seconds())
        if sec < 60:   return f"{sec}s ago"
        if sec < 3600: return f"{sec//60}m ago"
        return f"{sec//3600}h {(sec%3600)//60}m ago"
    except Exception:
        return "unknown"


def main():
    print()
    print("=" * 55)
    print("  CLEARING MONITOR – STATUS CHECK")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 55)

    # ── Heartbeat ──────────────────────────────────────────
    hb_path = cfg.BASE_DIR / "heartbeat.json"
    if not hb_path.exists():
        print("\n  STATUS : NOT STARTED")
        print("  The monitor has never run on this machine.")
        print("  Double-click run_monitor.bat to start it.")
        print()
        sys.exit(1)

    hb = json.loads(hb_path.read_text(encoding="utf-8"))
    last_poll_str = hb.get("last_poll_at", "")
    poll_num      = hb.get("poll_number", 0)
    interval_min  = hb.get("poll_interval_min", cfg.POLL_INTERVAL_MINUTES)

    try:
        last_poll_dt = datetime.fromisoformat(last_poll_str)
        age_sec      = int((datetime.now() - last_poll_dt).total_seconds())
        age_min      = age_sec / 60
        # Warn if last poll was more than 2× the interval ago
        if age_min <= interval_min * 2:
            status_line = f"  OK – running normally"
        else:
            status_line = f"  WARNING – last poll was {int(age_min)}m ago (expected every {interval_min}m)"
    except Exception:
        age_min    = None
        status_line = "  UNKNOWN – could not parse heartbeat timestamp"

    print(f"\n  Status      : {status_line}")
    print(f"  Last poll   : {last_poll_str}  ({_age(last_poll_str)})")
    print(f"  Poll count  : {poll_num}")
    print(f"  Interval    : every {interval_min} minutes")

    # Expected polls: how many should have run since the log started?
    log_files = sorted(cfg.LOGS_DIR.glob("monitor_*.log"))
    if log_files and age_min is not None:
        first_log   = log_files[0]
        started_at  = datetime.fromtimestamp(first_log.stat().st_ctime)
        elapsed_min = (datetime.now() - started_at).total_seconds() / 60
        expected    = max(1, int(elapsed_min / interval_min))
        missed      = max(0, expected - poll_num)
        print(f"  Running for : ~{int(elapsed_min)}m  (expected ~{expected} polls, ran {poll_num})")
        if missed == 0:
            print("  Missed polls: none")
        else:
            print(f"  Missed polls: ~{missed}  << check the logs folder")

    # ── Excel summary ──────────────────────────────────────
    print()
    if cfg.EXCEL_FILE.exists():
        try:
            from openpyxl import load_workbook
            wb   = load_workbook(cfg.EXCEL_FILE, read_only=True, data_only=True)
            print("  Excel file  :", cfg.EXCEL_FILE.name)
            for sheet_name in wb.sheetnames:
                ws          = wb[sheet_name]
                data_rows   = max(0, ws.max_row - 1) if ws.max_row else 0
                # Count non-empty rows (excluding headers)
                change_rows = sum(
                    1 for r in ws.iter_rows(min_row=2, values_only=True)
                    if any(v for v in r)
                )
                print(f"    Sheet [{sheet_name}] : {change_rows} change events recorded")
            wb.close()
        except Exception as e:
            print(f"  Excel       : could not read ({e})")
    else:
        print("  Excel       : not created yet (no changes detected so far)")

    # ── Recent log errors ──────────────────────────────────
    print()
    if log_files:
        latest_log = log_files[-1]
        lines      = latest_log.read_text(encoding="utf-8", errors="replace").splitlines()
        errors     = [l for l in lines if "ERROR" in l]
        warnings   = [l for l in lines if "WARNING" in l]
        print(f"  Log file    : {latest_log.name}  ({len(lines)} lines)")
        if errors:
            print(f"  Errors      : {len(errors)} found – last 2:")
            for e in errors[-2:]:
                print(f"    {e.strip()}")
        else:
            print("  Errors      : none")
        if warnings:
            print(f"  Warnings    : {len(warnings)}")
        # Show last 3 poll lines
        poll_lines = [l for l in lines if "Poll #" in l or "change(s) written" in l or "No changes" in l]
        if poll_lines:
            print("\n  Recent activity:")
            for l in poll_lines[-5:]:
                print(f"    {l.strip()}")
    else:
        print("  Logs        : none found")

    print()
    print("=" * 55)
    print()


if __name__ == "__main__":
    main()
