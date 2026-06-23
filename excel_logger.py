"""
excel_logger.py  –  per-course change tables inside each university sheet.

Sheet layout (one sheet per university):
  Each course gets its OWN small table:

    Row N   │ Course Name          │ 14-Aug 09:14  │ 14-Aug 11:02  │ 14-Aug 14:55  │
    Row N+1 │ Accounting & Finance │ ABB · 120 pts │ BBC · 112 pts │ BCC · 96 pts  │
    Row N+2 │ (blank separator)    │               │               │               │

Rules:
  • A new column is added ONLY when the entry requirement actually changes.
  • The first column shows the full requirement as first observed.
  • Every subsequent column also shows the full new requirement (exact text scraped).
  • No direction arrows, no colour-coded direction – just the facts and the time.
  • Timestamps are formatted "dd-Mon HH:MM" (e.g. "14-Aug 09:14").
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colours / styles
# ---------------------------------------------------------------------------
HEADER_BG   = "FF7B2D8B"   # Aston purple  – timestamp header cells
HEADER_FG   = "FFFFFFFF"   # white text
LABEL_BG    = "FF5A1F6E"   # darker purple – "Course Name" label cell
DATA_BG     = "FFFFFFFF"   # white         – requirement cells
COURSE_BG   = "FFF3E8F7"   # light purple  – course name cell

COL_A_WIDTH   = 44          # course name column
TIMESTAMP_WIDTH = 20        # each timestamp column


def _fill(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)


def _font(bold=False, colour="FF000000", size=10) -> Font:
    return Font(bold=bold, color=colour, size=size)


def _border() -> Border:
    t = Side(style="thin", color="FFAAAAAA")
    return Border(left=t, right=t, top=t, bottom=t)


def _centre() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _fmt_ts(dt: datetime) -> str:
    return dt.strftime("%d-%b %H:%M")


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _open_or_create(path: Path) -> Workbook:
    if path.exists():
        try:
            return load_workbook(path)
        except Exception as e:
            log.warning("Could not open workbook (%s) – creating fresh.", e)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _ensure_sheet(wb: Workbook, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions["A"].width = COL_A_WIDTH
        ws.sheet_view.showGridLines = True
    return wb[sheet_name]


# ---------------------------------------------------------------------------
# Course-table helpers
# ---------------------------------------------------------------------------

def _find_course_data_row(ws, course_name: str) -> int | None:
    """Return the data row (not header row) for this course, or None."""
    target = course_name.strip().lower()
    for row_cells in ws.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row_cells[0]
        if cell.value and str(cell.value).strip().lower() == target:
            return cell.row
    return None


def _next_empty_col(ws, header_row: int) -> int:
    """First column index (1-based) in header_row that has no value."""
    col = 2
    while ws.cell(header_row, col).value is not None:
        col += 1
    return col


def _first_data_start_row(ws) -> int:
    """Row where the first new course table should begin (after existing content)."""
    max_r = ws.max_row
    if max_r is None or max_r < 1:
        return 1
    # Find the last non-empty row
    last_used = 1
    for r in range(max_r, 0, -1):
        if any(ws.cell(r, c).value for c in range(1, ws.max_column + 2)):
            last_used = r
            break
    return last_used + 2   # one blank separator then start new table


def _write_header_cell(ws, row: int, col: int, text: str):
    """Write a purple header cell (timestamp or 'Course Name' label)."""
    c = ws.cell(row, col, text)
    c.font      = _font(bold=True, colour=HEADER_FG, size=10)
    c.fill      = _fill(HEADER_BG)
    c.alignment = _centre()
    c.border    = _border()
    if col > 1:
        ws.column_dimensions[get_column_letter(col)].width = TIMESTAMP_WIDTH


def _write_label_cell(ws, row: int):
    """Write the 'Course Name' label in col A of the header row."""
    c = ws.cell(row, 1, "Course Name")
    c.font      = _font(bold=True, colour=HEADER_FG, size=10)
    c.fill      = _fill(LABEL_BG)
    c.alignment = _left()
    c.border    = _border()


def _write_course_name_cell(ws, row: int, course_name: str):
    """Write the actual course name in col A of the data row."""
    c = ws.cell(row, 1, course_name)
    c.font      = _font(bold=True, colour="FF3D0066", size=10)
    c.fill      = _fill(COURSE_BG)
    c.alignment = _left()
    c.border    = _border()


def _write_req_cell(ws, row: int, col: int, req: str):
    """Write an entry requirement value cell."""
    c = ws.cell(row, col, req)
    c.font      = _font(bold=False, colour="FF000000", size=10)
    c.fill      = _fill(DATA_BG)
    c.alignment = _centre()
    c.border    = _border()
    ws.row_dimensions[row].height = 30


def _create_course_table(ws, course_name: str, timestamp: datetime, entry_req: str):
    """Append a brand-new 2-row course table to the sheet."""
    header_row = _first_data_start_row(ws)
    data_row   = header_row + 1

    _write_label_cell(ws, header_row)
    _write_header_cell(ws, header_row, 2, _fmt_ts(timestamp))
    _write_course_name_cell(ws, data_row, course_name)
    _write_req_cell(ws, data_row, 2, entry_req)

    ws.row_dimensions[header_row].height = 22
    ws.row_dimensions[data_row].height   = 36

    return header_row, data_row


def _append_column(ws, header_row: int, data_row: int, timestamp: datetime, entry_req: str):
    """Add a new timestamp+requirement column to an existing course table."""
    col = _next_empty_col(ws, header_row)
    _write_header_cell(ws, header_row, col, _fmt_ts(timestamp))
    _write_req_cell(ws, data_row, col, entry_req)
    return col


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def log_changes(path: Path, uni_key: str, changes: list[dict], data_dir):
    """
    For each change event, either create a new course table or append a column
    to the existing one. Only called when at least one change was detected.
    """
    if not changes:
        return

    wb = _open_or_create(path)
    ws = _ensure_sheet(wb, uni_key[:31])

    for ch in changes:
        course    = ch["course"]
        new_req   = ch["new_req"]
        detected  = ch["detected_at"]

        # Skip "removed" events – course gone from clearing, no new req to show
        if new_req == "REMOVED":
            continue

        data_row = _find_course_data_row(ws, course)

        if data_row is None:
            # Brand-new course in this sheet – create its table
            _create_course_table(ws, course, detected, new_req)
        else:
            # Existing course – append a new timestamp column
            header_row = data_row - 1
            _append_column(ws, header_row, data_row, detected, new_req)

    _safe_save(wb, path)
    log.info("Excel updated -> %s  (%d changes recorded)", path.name, len(changes))


def init_workbook(path: Path, universities: dict):
    """Create the workbook with a sheet per university if it doesn't exist."""
    wb = _open_or_create(path)
    for uni_key in universities:
        _ensure_sheet(wb, uni_key[:31])
    _safe_save(wb, path)


def _safe_save(wb: Workbook, path: Path):
    try:
        wb.save(path)
    except PermissionError:
        alt = path.with_stem(path.stem + "_" + datetime.now().strftime("%H%M%S"))
        wb.save(alt)
        log.warning(
            "%s is open in Excel – saved to %s instead. "
            "Close Excel before the next run.", path.name, alt.name
        )
